"""Check the source data loading."""
import openpyxl
from pathlib import Path

def check_source_data():
    """Print the contents of the source Excel file."""
    data_dir = Path(__file__).parent / "data"
    source_file = data_dir / "source.xlsx"
    
    print("📊 Checking source file contents...")
    print("=" * 60)
    
    wb = openpyxl.load_workbook(source_file, read_only=True)
    
    # Check CustomerData sheet
    if "CustomerData" in wb.sheetnames:
        ws = wb["CustomerData"]
        print("\n🧑 CustomerData Sheet:")
        print("-" * 40)
        
        # Print headers
        headers = [str(cell.value) for cell in ws[1] if cell.value]
        print("Headers:", ", ".join(headers))
        
        # Print first few rows
        print("\nSample Data (first 3 rows):")
        rows = list(ws.rows)
        if len(rows) > 1:
            for row in rows[1:4]:  # Skip header, take next 3 rows
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                print(" | ".join(row_data))
        else:
            print("No data rows found!")
    
    # Check Transactions sheet
    if "Transactions" in wb.sheetnames:
        ws = wb["Transactions"]
        print("\n💰 Transactions Sheet:")
        print("-" * 40)
        
        # Print headers
        headers = [str(cell.value) for cell in ws[1] if cell.value]
        print("Headers:", ", ".join(headers))
        
        # Print first few rows
        print("\nSample Data (first 3 rows):")
        rows = list(ws.rows)
        if len(rows) > 1:
            for row in rows[1:4]:  # Skip header, take next 3 rows
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                print(" | ".join(row_data))
        else:
            print("No data rows found!")
    
    wb.close()
    print("\n" + "=" * 60)

if __name__ == "__main__":
    check_source_data()