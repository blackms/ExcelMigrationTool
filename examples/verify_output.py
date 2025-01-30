"""Verify the output Excel file contents."""
import openpyxl
from pathlib import Path

def verify_output():
    """Print the contents of the output Excel file."""
    data_dir = Path(__file__).parent / "data"
    output_file = data_dir / "test_output.xlsx"
    
    print("📊 Verifying output file contents...")
    print("=" * 60)
    
    if not output_file.exists():
        print("\n❌ Output file not found!")
        return
    
    wb = openpyxl.load_workbook(output_file, read_only=True)
    
    # Check CustomerSummary sheet
    if "CustomerSummary" in wb.sheetnames:
        ws = wb["CustomerSummary"]
        print("\n🧑 CustomerSummary Sheet:")
        print("-" * 40)
        
        # Get all rows
        rows = list(ws.rows)
        if not rows:
            print("Sheet is empty!")
            return
        
        # Print headers
        headers = [str(cell.value) for cell in rows[0] if cell.value is not None]
        if headers:
            print("Headers:", ", ".join(headers))
        else:
            print("No headers found!")
        
        # Print first few rows
        if len(rows) > 1:
            print("\nSample Data (first 3 rows):")
            for row in rows[1:min(4, len(rows))]:  # Skip header, take next 3 rows
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                print(" | ".join(row_data))
        else:
            print("No data rows found!")
    
    # Check TransactionSummary sheet
    if "TransactionSummary" in wb.sheetnames:
        ws = wb["TransactionSummary"]
        print("\n💰 TransactionSummary Sheet:")
        print("-" * 40)
        
        # Get all rows
        rows = list(ws.rows)
        if not rows:
            print("Sheet is empty!")
            return
        
        # Print headers
        headers = [str(cell.value) for cell in rows[0] if cell.value is not None]
        if headers:
            print("Headers:", ", ".join(headers))
        else:
            print("No headers found!")
        
        # Print first few rows
        if len(rows) > 1:
            print("\nSample Data (first 3 rows):")
            for row in rows[1:min(4, len(rows))]:  # Skip header, take next 3 rows
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                print(" | ".join(row_data))
        else:
            print("No data rows found!")
    
    wb.close()
    print("\n" + "=" * 60)

if __name__ == "__main__":
    verify_output()