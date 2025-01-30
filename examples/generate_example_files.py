"""Generate example Excel files for demonstrating rule generation."""
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
import datetime
from decimal import Decimal

def create_source_file(filepath: Path):
    """Create source Excel file with sample data."""
    wb = openpyxl.Workbook()
    
    # Customer Data Sheet
    ws = wb.active
    ws.title = "CustomerData"
    
    # Headers
    headers = [
        "CustomerID", "FirstName", "LastName", "Email", 
        "PhoneNumber", "RegistrationDate", "TotalPurchases"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Sample data
    data = [
        [1001, "John", "Doe", "john.doe@email.com", "123-456-7890", 
         datetime.date(2023, 1, 15), Decimal("1250.50")],
        [1002, "Jane", "Smith", "jane.smith@email.com", "234-567-8901",
         datetime.date(2023, 2, 20), Decimal("2100.75")],
        [1003, "Bob", "Johnson", "bob.j@email.com", "345-678-9012",
         datetime.date(2023, 3, 10), Decimal("750.25")],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Transaction Sheet
    ws = wb.create_sheet("Transactions")
    
    # Headers
    headers = [
        "TransactionID", "CustomerID", "Date", "ProductID",
        "Quantity", "UnitPrice", "Total"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Sample data
    data = [
        ["T001", 1001, datetime.date(2023, 6, 1), "P101", 2, Decimal("25.99"), Decimal("51.98")],
        ["T002", 1001, datetime.date(2023, 6, 15), "P102", 1, Decimal("99.99"), Decimal("99.99")],
        ["T003", 1002, datetime.date(2023, 6, 20), "P101", 3, Decimal("25.99"), Decimal("77.97")],
        ["T004", 1003, datetime.date(2023, 6, 25), "P103", 1, Decimal("149.99"), Decimal("149.99")],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(filepath)

def create_target_file(filepath: Path):
    """Create target Excel file showing desired transformations."""
    wb = openpyxl.Workbook()
    
    # Customer Summary Sheet
    ws = wb.active
    ws.title = "CustomerSummary"
    
    # Headers
    headers = [
        "CustomerID", "FullName", "ContactInfo", 
        "MemberSince", "PurchaseMetrics", "Status"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Sample transformed data
    data = [
        [1001, "John Doe", "Email: john.doe@email.com\nPhone: 123-456-7890",
         "January 2023", "Total: $1,250.50\nTransactions: 2", "Active"],
        [1002, "Jane Smith", "Email: jane.smith@email.com\nPhone: 234-567-8901",
         "February 2023", "Total: $2,100.75\nTransactions: 1", "Active"],
        [1003, "Bob Johnson", "Email: bob.j@email.com\nPhone: 345-678-9012",
         "March 2023", "Total: $750.25\nTransactions: 1", "Active"],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 3:  # ContactInfo column
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            if col_idx == 5:  # PurchaseMetrics column
                cell.alignment = openpyxl.styles.Alignment(wrapText=True)
    
    # Transaction Summary Sheet
    ws = wb.create_sheet("TransactionSummary")
    
    # Headers
    headers = [
        "TransactionID", "Customer", "Date", "Product",
        "OrderDetails", "TotalAmount"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        cell.font = Font(bold=True)
    
    # Sample transformed data
    data = [
        ["T001", "John Doe", "Jun 1, 2023", "P101", "2 units @ $25.99", "$51.98"],
        ["T002", "John Doe", "Jun 15, 2023", "P102", "1 unit @ $99.99", "$99.99"],
        ["T003", "Jane Smith", "Jun 20, 2023", "P101", "3 units @ $25.99", "$77.97"],
        ["T004", "Bob Johnson", "Jun 25, 2023", "P103", "1 unit @ $149.99", "$149.99"],
    ]
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    wb.save(filepath)

def main():
    """Generate example files."""
    # Create data directory if it doesn't exist
    data_dir = Path(__file__).parent / "data"
    data_dir.mkdir(exist_ok=True)
    
    # Generate files
    create_source_file(data_dir / "source.xlsx")
    create_target_file(data_dir / "target.xlsx")
    
    print("✨ Example files generated successfully!")
    print(f"📊 Source file: {data_dir / 'source.xlsx'}")
    print(f"🎯 Target file: {data_dir / 'target.xlsx'}")

if __name__ == "__main__":
    main()