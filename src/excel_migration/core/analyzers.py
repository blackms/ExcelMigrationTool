"""Concrete implementations of analyzer interfaces."""
from pathlib import Path
from typing import Dict, Any
import openpyxl
from loguru import logger

from .interfaces import SheetAnalyzer
from ..vision.processor import SheetImageProcessor

class ExcelSheetAnalyzer(SheetAnalyzer):
    """Concrete implementation of sheet analyzer."""
    
    def __init__(self, image_processor: SheetImageProcessor):
        self.image_processor = image_processor
        logger.debug("Initialized Excel sheet analyzer")

    async def analyze_sheet(self, sheet_path: Path, sheet_name: str) -> Dict[str, Any]:
        """Analyze a sheet and return its structure and content."""
        try:
            # First pass with read_only for basic data
            analysis = self._analyze_data(sheet_path, sheet_name)
            
            # Second pass without read_only for formulas
            analysis.update(self._analyze_formulas(sheet_path, sheet_name))
            
            return analysis
            
        except Exception as e:
            logger.error(f"Failed to analyze sheet {sheet_name} in {sheet_path}: {str(e)}")
            raise

    def _analyze_data(self, sheet_path: Path, sheet_name: str) -> Dict[str, Any]:
        """Analyze sheet data using read_only mode."""
        wb = openpyxl.load_workbook(sheet_path, read_only=True)
        ws = wb[sheet_name]
        
        analysis = {
            "sheet_name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "headers": [],
            "data_sample": [],
            "column_types": {}
        }
        
        # Get headers (first row)
        for cell in ws[1]:
            if cell.value:
                analysis["headers"].append(str(cell.value))
        
        # Sample some data rows
        sample_size = min(5, ws.max_row - 1)  # Sample up to 5 rows
        for row in list(ws.rows)[1:sample_size+1]:  # Skip header row
            row_data = []
            for cell in row:
                if cell.value:
                    row_data.append(str(cell.value))
                else:
                    row_data.append("")
            analysis["data_sample"].append(row_data)
        
        # Analyze column types
        for col in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            values = []
            for row in range(2, min(7, ws.max_row + 1)):  # Sample first 5 data rows
                cell = ws[f"{col_letter}{row}"]
                if cell.value:
                    values.append(type(cell.value).__name__)
            
            # Determine most common type
            if values:
                from collections import Counter
                type_counts = Counter(values)
                most_common_type = type_counts.most_common(1)[0][0]
                analysis["column_types"][col_letter] = most_common_type
        
        wb.close()
        return analysis

    def _analyze_formulas(self, sheet_path: Path, sheet_name: str) -> Dict[str, Any]:
        """Analyze sheet formulas without read_only mode."""
        try:
            wb = openpyxl.load_workbook(sheet_path, data_only=False)
            ws = wb[sheet_name]
            
            formulas = []
            # Only check first 100 rows to avoid performance issues
            max_rows = min(100, ws.max_row)
            
            for row in range(1, max_rows + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if hasattr(cell, 'value') and cell.value and str(cell.value).startswith('='):
                        formulas.append({
                            "cell": f"{openpyxl.utils.get_column_letter(col)}{row}",
                            "formula": str(cell.value)[1:]  # Remove the '=' prefix
                        })
            
            wb.close()
            return {"formulas": formulas}
            
        except Exception as e:
            logger.warning(f"Could not analyze formulas in {sheet_name}: {str(e)}")
            return {"formulas": []}  # Return empty formulas on error