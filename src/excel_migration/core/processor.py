"""Core Excel migration processor."""
from typing import Optional, Dict, Any, List
import openpyxl
from pathlib import Path
import logging

from .models import (
    MigrationContext,
    MigrationRule,
    Cell,
    CellType,
    ValidationResult
)

logger = logging.getLogger(__name__)

class ExcelMigrationProcessor:
    """Main processor for Excel migrations."""
    
    def __init__(self, context: MigrationContext):
        """Initialize with migration context."""
        self.context = context
        self.source_wb = None
        self.target_wb = None
        self._setup_logging()

    def _setup_logging(self):
        """Configure logging."""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )

    def process(self) -> bool:
        """Execute the migration process."""
        try:
            # Load workbooks
            self.source_wb = self._load_workbook(self.context.source_file)
            self.target_wb = openpyxl.Workbook()

            # Process each sheet mapping
            for source_sheet_name, target_sheet_name in self.context.sheet_mapping.items():
                success = self._process_sheet(source_sheet_name, target_sheet_name)
                if not success:
                    return False

            # Save target workbook
            self.target_wb.save(self.context.target_file)
            return True

        except Exception as e:
            logger.error(f"Migration failed: {str(e)}")
            return False

        finally:
            self._cleanup()

    def _load_workbook(self, file_path: str) -> openpyxl.Workbook:
        """Load an Excel workbook."""
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        return openpyxl.load_workbook(file_path, data_only=False)

    def _process_sheet(self, source_sheet_name: str, target_sheet_name: str) -> bool:
        """Process a single sheet according to rules."""
        try:
            source_sheet = self.source_wb[source_sheet_name]
            target_sheet = self.target_wb.create_sheet(target_sheet_name)

            # Get applicable rules for this sheet
            sheet_rules = [rule for rule in self.context.rules 
                         if rule.source_columns[0].split('!')[0] == source_sheet_name]

            # Process each row
            header_row = self._find_header_row(source_sheet)
            for row in range(header_row + 1, source_sheet.max_row + 1):
                self._process_row(row, source_sheet, target_sheet, sheet_rules)

            return True

        except Exception as e:
            logger.error(f"Failed to process sheet {source_sheet_name}: {str(e)}")
            return False

    def _find_header_row(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        """Find the header row in a sheet."""
        for row in range(1, min(10, sheet.max_row + 1)):
            if any(cell.value for cell in sheet[row]):
                return row
        return 1

    def _process_row(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                    target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                    rules: List[MigrationRule]) -> None:
        """Process a single row according to rules."""
        for rule in rules:
            # Extract source values
            source_values = self._get_source_values(row, source_sheet, rule)
            
            # Apply rule
            result = self._apply_rule(rule, source_values)
            
            # Write result to target
            if result is not None:
                self._write_result(row, target_sheet, rule.target_column, result)

    def _get_source_values(self, row: int, sheet: openpyxl.worksheet.worksheet.Worksheet,
                         rule: MigrationRule) -> Dict[str, Cell]:
        """Get source values for a rule."""
        values = {}
        for col_ref in rule.source_columns:
            sheet_name, col = col_ref.split('!') if '!' in col_ref else ('', col_ref)
            if not sheet_name or sheet_name == sheet.title:
                cell = sheet[f"{col}{row}"]
                values[col] = Cell(
                    value=cell.value,
                    cell_type=self._determine_cell_type(cell),
                    row=row,
                    column=openpyxl.utils.column_index_from_string(col),
                    formula=cell.formula if cell.formula else None,
                    style={
                        'font': cell.font,
                        'fill': cell.fill,
                        'border': cell.border,
                        'alignment': cell.alignment,
                        'number_format': cell.number_format
                    }
                )
        return values

    def _determine_cell_type(self, cell: openpyxl.cell.cell.Cell) -> CellType:
        """Determine the type of a cell."""
        if cell.formula:
            return CellType.FORMULA
        if isinstance(cell.value, (int, float)):
            return CellType.NUMBER
        if isinstance(cell.value, bool):
            return CellType.BOOLEAN
        if isinstance(cell.value, str):
            return CellType.TEXT
        return CellType.TEXT

    def _apply_rule(self, rule: MigrationRule, source_values: Dict[str, Cell]) -> Any:
        """Apply a migration rule to source values."""
        # Implement rule application logic based on rule type
        # This will be extended with LLM integration for complex transformations
        pass

    def _write_result(self, row: int, sheet: openpyxl.worksheet.worksheet.Worksheet,
                     column: str, value: Any) -> None:
        """Write a result to the target sheet."""
        cell = sheet[f"{column}{row}"]
        cell.value = value

    def _cleanup(self) -> None:
        """Clean up resources."""
        if self.source_wb:
            self.source_wb.close()
        if self.target_wb:
            self.target_wb.close()