"""Rule generation and execution engine."""
from pathlib import Path
from typing import Dict, Any, List, Optional
import openpyxl
from loguru import logger
from langchain_openai import ChatOpenAI

class RuleEngine:
    """Engine for generating and executing migration rules."""
    
    def __init__(self, llm_provider: str = "openai"):
        """Initialize the rule engine."""
        self.llm = ChatOpenAI(
            model_name="gpt-4",
            temperature=0.7
        )
        logger.debug(f"Initialized rule engine with {llm_provider}")

    async def generate_rules(
        self,
        source_file: Path,
        target_file: Path,
        source_sheet: str,
        target_sheet: str
    ) -> List[Dict[str, Any]]:
        """Generate migration rules by analyzing example files."""
        try:
            # Analyze source and target structures
            source_structure = self._analyze_sheet(source_file, source_sheet)
            target_structure = self._analyze_sheet(target_file, target_sheet)
            
            rules = []
            
            # Direct field mappings
            direct_mappings = self._generate_direct_mappings(
                source_structure,
                target_structure
            )
            rules.extend(direct_mappings)
            
            # Transformation rules
            transform_rules = self._generate_transformation_rules(
                source_structure,
                target_structure
            )
            rules.extend(transform_rules)
            
            # Calculation rules
            calc_rules = self._generate_calculation_rules(
                source_structure,
                target_structure
            )
            rules.extend(calc_rules)
            
            # Remove duplicates
            unique_rules = []
            seen_targets = set()
            for rule in rules:
                target = rule["target_field"]
                if target not in seen_targets:
                    unique_rules.append(rule)
                    seen_targets.add(target)
            
            return unique_rules
            
        except Exception as e:
            logger.error(f"Rule generation failed: {str(e)}")
            return []

    def _analyze_sheet(self, file_path: Path, sheet_name: str) -> Dict[str, Any]:
        """Analyze sheet structure and content."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]
            
            analysis = {
                "sheet_name": sheet_name,
                "headers": [],
                "data_types": {},
                "sample_data": []
            }
            
            # Get headers
            for cell in ws[1]:
                if cell.value:
                    analysis["headers"].append(str(cell.value))
            
            # Analyze data types and get samples
            for col_idx, header in enumerate(analysis["headers"], 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                values = []
                for row in range(2, min(7, ws.max_row + 1)):  # Sample first 5 data rows
                    cell = ws[f"{col_letter}{row}"]
                    if cell.value:
                        values.append(str(cell.value))
                
                if values:
                    analysis["data_types"][header] = self._infer_data_type(values)
                    analysis["sample_data"].append({
                        "header": header,
                        "samples": values
                    })
            
            wb.close()
            return analysis
            
        except Exception as e:
            logger.error(f"Sheet analysis failed: {str(e)}")
            raise

    def _infer_data_type(self, values: List[str]) -> str:
        """Infer data type from sample values."""
        try:
            # Try numeric
            float(values[0])
            return "numeric"
        except ValueError:
            pass
        
        # Check date format
        date_indicators = ["/", "-", ":", "AM", "PM"]
        if any(ind in values[0] for ind in date_indicators):
            return "datetime"
        
        # Check boolean
        bool_values = ["true", "false", "yes", "no", "0", "1"]
        if all(v.lower() in bool_values for v in values):
            return "boolean"
        
        return "text"

    def _generate_direct_mappings(
        self,
        source: Dict[str, Any],
        target: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Generate direct field mapping rules."""
        rules = []
        
        # Direct matches
        for target_field in target["headers"]:
            if target_field in source["headers"]:
                rules.append({
                    "type": "field_mapping",
                    "source_field": target_field,
                    "target_field": target_field,
                    "transformation": self._get_transformation_rule(
                        source["data_types"].get(target_field),
                        target_field
                    )
                })
        
        return rules

    def _generate_transformation_rules(
        self,
        source: Dict[str, Any],
        target: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Generate transformation rules."""
        rules = []
        
        # Name transformations
        if "FullName" in target["headers"] and "FirstName" in source["headers"] and "LastName" in source["headers"]:
            rules.append({
                "type": "field_mapping",
                "source_field": ["FirstName", "LastName"],
                "target_field": "FullName",
                "transformation": {
                    "type": "concatenate",
                    "params": {
                        "separator": " "
                    }
                }
            })
        
        # Status transformations
        if "IsActive" in target["headers"] and "Status" in source["headers"]:
            rules.append({
                "type": "field_mapping",
                "source_field": "Status",
                "target_field": "IsActive",
                "transformation": {
                    "type": "boolean_transform",
                    "params": {
                        "true_values": ["Active"],
                        "false_values": ["Inactive"]
                    }
                }
            })
        
        return rules

    def _generate_calculation_rules(
        self,
        source: Dict[str, Any],
        target: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Generate calculation rules."""
        rules = []
        
        # Days since registration
        if "DaysSinceRegistration" in target["headers"] and "RegistrationDate" in source["headers"]:
            rules.append({
                "type": "calculation",
                "target_field": "DaysSinceRegistration",
                "formula": "DATEDIF([RegistrationDate], TODAY(), 'D')",
                "description": "Calculate days between registration date and today",
                "source_fields": ["RegistrationDate"]
            })
        
        # Transaction count
        if "TransactionCount" in target["headers"] and "TransactionID" in source["headers"]:
            rules.append({
                "type": "calculation",
                "target_field": "TransactionCount",
                "formula": "COUNT([TransactionID])",
                "description": "Count total number of transactions",
                "source_fields": ["TransactionID"]
            })
        
        # Total spent
        if "TotalSpent" in target["headers"] and "Amount" in source["headers"]:
            rules.append({
                "type": "calculation",
                "target_field": "TotalSpent",
                "formula": "SUM([Amount])",
                "description": "Sum of all transaction amounts",
                "source_fields": ["Amount"]
            })
        
        # Average amount
        if "AverageAmount" in target["headers"] and "Amount" in source["headers"]:
            rules.append({
                "type": "calculation",
                "target_field": "AverageAmount",
                "formula": "AVERAGE([Amount])",
                "description": "Average transaction amount",
                "source_fields": ["Amount"]
            })
        
        # Success rate
        if "SuccessRate" in target["headers"] and "Status" in source["headers"]:
            rules.append({
                "type": "calculation",
                "target_field": "SuccessRate",
                "formula": "COUNT_IF([Status], 'Completed') / COUNT([Status])",
                "description": "Percentage of completed transactions",
                "source_fields": ["Status"]
            })
        
        return rules

    def _get_transformation_rule(self, source_type: str, target_field: str) -> Dict[str, Any]:
        """Get appropriate transformation rule based on field types."""
        transformation = {
            "type": "direct",  # Default to direct copy
            "params": {}
        }
        
        # Add type-specific transformations
        if source_type == "datetime":
            transformation.update({
                "type": "datetime_format",
                "params": {
                    "format": self._infer_date_format(target_field)
                }
            })
        elif source_type == "numeric":
            transformation.update({
                "type": "numeric_format",
                "params": {
                    "decimal_places": 2 if "amount" in target_field.lower() else 0,
                    "thousands_separator": True
                }
            })
        
        return transformation

    def _infer_date_format(self, field_name: str) -> str:
        """Infer date format based on field name."""
        if any(word in field_name.lower() for word in ["time", "timestamp"]):
            return "%Y-%m-%d %H:%M:%S"
        return "%Y-%m-%d"