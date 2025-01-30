"""Base task implementations for Excel migration framework."""
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set
from loguru import logger
import openpyxl

from ..core.interfaces import (
    Task, TaskHandler, TaskProcessor, RuleGenerator,
    SheetAnalyzer, DataExtractor, ImageProcessor
)

@dataclass
class SheetMapping:
    """Mapping between source and target sheets."""
    source_sheet: str
    target_sheet: str
    rules: Optional[List[Dict[str, Any]]] = None
    screenshot: Optional[Path] = None
    validation_errors: List[str] = field(default_factory=list)

    def __post_init__(self):
        """Validate the mapping."""
        if not self.source_sheet or not self.target_sheet:
            raise ValueError("Source and target sheet names are required")

@dataclass
class MigrationTask:
    """Base migration task implementation."""
    source_file: Path
    target_file: Path
    task_type: str
    description: str
    context: Dict[str, Any]
    sheet_mappings: List[SheetMapping] = field(default_factory=list)
    example_source: Optional[Path] = None
    example_target: Optional[Path] = None
    example_sheet_mappings: List[SheetMapping] = field(default_factory=list)
    screenshots: Dict[Path, str] = field(default_factory=dict)  # screenshot path -> sheet name

    def __post_init__(self):
        """Initialize task with defaults and validate."""
        self.context = self.context or {}
        self._validate_files()
        self._validate_sheets()
        logger.info(f"Initialized {self.task_type} task: {self.description}")

    def _validate_files(self):
        """Validate file existence and format."""
        if not self.source_file.exists():
            raise FileNotFoundError(f"Source file not found: {self.source_file}")
        
        if self.example_source and not self.example_source.exists():
            raise FileNotFoundError(f"Example source file not found: {self.example_source}")
        
        if self.example_target and not self.example_target.exists():
            raise FileNotFoundError(f"Example target file not found: {self.example_target}")
        
        for screenshot in self.screenshots:
            if not screenshot.exists():
                raise FileNotFoundError(f"Screenshot not found: {screenshot}")

    def _validate_sheets(self):
        """Validate sheet existence and mappings."""
        try:
            # Check source sheets
            wb = openpyxl.load_workbook(self.source_file, read_only=True)
            source_sheets = set(wb.sheetnames)
            wb.close()

            for mapping in self.sheet_mappings:
                if mapping.source_sheet not in source_sheets:
                    raise ValueError(f"Sheet not found in source file: {mapping.source_sheet}")

            # Check example sheets if provided
            if self.example_source and self.example_target:
                wb = openpyxl.load_workbook(self.example_source, read_only=True)
                example_source_sheets = set(wb.sheetnames)
                wb.close()

                wb = openpyxl.load_workbook(self.example_target, read_only=True)
                example_target_sheets = set(wb.sheetnames)
                wb.close()
                
                for mapping in self.example_sheet_mappings:
                    if mapping.source_sheet not in example_source_sheets:
                        raise ValueError(f"Sheet not found in example source: {mapping.source_sheet}")
                    if mapping.target_sheet not in example_target_sheets:
                        raise ValueError(f"Sheet not found in example target: {mapping.target_sheet}")

        except Exception as e:
            logger.error(f"Sheet validation failed: {str(e)}")
            raise

class TaskRegistry:
    """Registry for task handlers."""
    
    def __init__(self):
        self._handlers: Dict[str, TaskHandler] = {}
        self._sheet_processors: Dict[str, TaskProcessor] = {}
    
    def register(self, task_type: str, handler: TaskHandler, sheet_processor: Optional[TaskProcessor] = None):
        """Register a handler and optional sheet processor for a task type."""
        self._handlers[task_type] = handler
        if sheet_processor:
            self._sheet_processors[task_type] = sheet_processor
        logger.debug(f"Registered handler for task type: {task_type}")
    
    async def get_handler(self, task: Task) -> Optional[TaskHandler]:
        """Get appropriate handler for a task."""
        for handler in self._handlers.values():
            if await handler.can_handle(task):
                return handler
        return None
    
    def get_sheet_processor(self, task_type: str) -> Optional[TaskProcessor]:
        """Get sheet processor for a task type."""
        return self._sheet_processors.get(task_type)

class BaseTaskHandler(TaskHandler):
    """Base implementation of task handler."""
    
    def __init__(self, processor: TaskProcessor):
        self.processor = processor
        logger.debug(f"Initialized {self.__class__.__name__}")
    
    async def can_handle(self, task: Task) -> bool:
        """Check if this handler can process the task."""
        return await self.processor.validate(task)
    
    async def handle(self, task: Task) -> bool:
        """Handle the task."""
        try:
            logger.info(f"Processing task: {task.description}")
            
            # Process each sheet mapping
            for mapping in task.sheet_mappings:
                success = await self._process_sheet_mapping(task, mapping)
                if not success:
                    return False
            
            return True
            
        except Exception as e:
            logger.exception(f"Task handling failed: {str(e)}")
            return False
    
    async def _process_sheet_mapping(self, task: Task, mapping: SheetMapping) -> bool:
        """Process a single sheet mapping."""
        try:
            logger.info(f"Processing sheet mapping: {mapping.source_sheet} -> {mapping.target_sheet}")
            
            # Apply screenshot analysis if available
            if mapping.screenshot:
                await self._analyze_screenshot(task, mapping)
            
            # Process the sheet
            return await self.processor.process_sheet(task, mapping)
            
        except Exception as e:
            logger.exception(f"Sheet mapping processing failed: {str(e)}")
            return False
    
    async def _analyze_screenshot(self, task: Task, mapping: SheetMapping):
        """Analyze screenshot for additional context."""
        try:
            if not mapping.screenshot:
                return
            
            image_processor = task.context.get("image_processor")
            if not image_processor:
                logger.warning("No image processor available for screenshot analysis")
                return
            
            analysis = await image_processor.process_image(mapping.screenshot)
            mapping.context = mapping.context or {}
            mapping.context["screenshot_analysis"] = analysis
            
        except Exception as e:
            logger.exception(f"Screenshot analysis failed: {str(e)}")

class TaskBasedProcessor(TaskProcessor):
    """Process tasks with rule generation and multimodal analysis."""
    
    def __init__(self, 
                 rule_generator: RuleGenerator,
                 sheet_analyzer: SheetAnalyzer,
                 llm_provider: Any):
        self.rule_generator = rule_generator
        self.sheet_analyzer = sheet_analyzer
        self.llm_provider = llm_provider
        logger.debug("Initialized task-based processor")
    
    async def process(self, task: Task) -> bool:
        """Process a task."""
        try:
            # Generate rules if example files are provided
            if task.example_source and task.example_target:
                await self._generate_rules_from_examples(task)
            
            # Process each sheet mapping
            for mapping in task.sheet_mappings:
                success = await self.process_sheet(task, mapping)
                if not success:
                    return False
            
            return True
            
        except Exception as e:
            logger.exception(f"Task processing failed: {str(e)}")
            return False
    
    async def process_sheet(self, task: Task, mapping: SheetMapping) -> bool:
        """Process a single sheet mapping."""
        try:
            # Load source data
            source_rows = self._load_sheet_data(task.source_file, mapping.source_sheet)
            
            # Process each row
            for source_data in source_rows:
                task.context["source_data"] = source_data
                task.context["target_data"] = {}
                
                # Analyze source sheet (only once per sheet)
                if "sheet_analysis" not in task.context:
                    analysis = await self.sheet_analyzer.analyze_sheet(
                        task.source_file,
                        mapping.source_sheet
                    )
                    task.context["sheet_analysis"] = analysis
                
                # Get insights from LLM (only once per sheet)
                if "sheet_insights" not in task.context:
                    insights = await self.llm_provider.analyze_task({
                        **task.context,
                        "sheet_analysis": task.context["sheet_analysis"],
                        "mapping": mapping
                    })
                    task.context["sheet_insights"] = insights
                
                # Apply rules to this row
                if mapping.rules:
                    for rule in mapping.rules:
                        success = await self._apply_rule(task, mapping, rule)
                        if not success:
                            return False
                
                # Save target data if in migration mode
                if task.task_type == "migrate" and task.context["target_data"]:
                    self._save_sheet_data(
                        task.target_file,
                        mapping.target_sheet,
                        task.context["target_data"]
                    )
            
            return True
            
        except Exception as e:
            logger.exception(f"Sheet processing failed: {str(e)}")
            return False
    
    def _load_sheet_data(self, file_path: Path, sheet_name: str) -> List[Dict[str, Any]]:
        """Load data from a sheet into a list of dictionaries."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            # Load data
            data_rows = []
            rows = list(ws.rows)[1:]  # Skip header row
            
            if sheet_name == "CustomerData":
                # For CustomerData, each row is a separate customer
                for row in rows:
                    row_data = {}
                    for header, cell in zip(headers, row):
                        if cell.value is not None:
                            row_data[header] = cell.value
                    if row_data:  # Only add non-empty rows
                        data_rows.append(row_data)
            
            elif sheet_name == "Transactions":
                # For Transactions, group by CustomerID
                customer_transactions = {}
                for row in rows:
                    row_data = {}
                    customer_id = None
                    for header, cell in zip(headers, row):
                        if cell.value is not None:
                            row_data[header] = cell.value
                            if header == "CustomerID":
                                customer_id = cell.value
                    
                    if customer_id and row_data:
                        if customer_id not in customer_transactions:
                            customer_transactions[customer_id] = []
                        customer_transactions[customer_id].append(row_data)
                
                # Convert grouped transactions to list format
                for customer_id, transactions in customer_transactions.items():
                    data_rows.append({
                        "CustomerID": customer_id,
                        "Transactions": transactions,
                        "TransactionCount": len(transactions),
                        "TotalAmount": sum(float(t["Amount"]) for t in transactions),
                        "AverageAmount": sum(float(t["Amount"]) for t in transactions) / len(transactions),
                        "LastTransactionDate": max(t["Date"] for t in transactions),
                        "SuccessRate": len([t for t in transactions if t["Status"] == "Completed"]) / len(transactions)
                    })
            
            wb.close()
            return data_rows
            
        except Exception as e:
            logger.error(f"Failed to load sheet data: {str(e)}")
            return []
    
    def _save_sheet_data(self, file_path: Path, sheet_name: str, data: Dict[str, Any]):
        """Save data to a sheet."""
        try:
            # Create new workbook if file doesn't exist
            if not file_path.exists():
                wb = openpyxl.Workbook()
                wb.remove(wb.active)  # Remove default sheet
            else:
                wb = openpyxl.load_workbook(file_path)
            
            # Create or get sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
                # Write headers for new sheet
                headers = list(data.keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            
            # Get current row count (excluding header)
            data_rows = max(0, ws.max_row - 1)
            
            # Write data row
            row_num = data_rows + 2  # Add 2 (1 for header, 1 for new row)
            for col, (header, value) in enumerate(data.items(), 1):
                ws.cell(row=row_num, column=col, value=value)
            
            # Save workbook
            wb.save(file_path)
            logger.debug(f"Saved row {data_rows + 1} to sheet {sheet_name} in {file_path}")
            
        except Exception as e:
            logger.error(f"Failed to save sheet data: {str(e)}")
    
    async def _generate_rules_from_examples(self, task: Task):
        """Generate rules from example files."""
        try:
            all_rules = []
            for mapping in task.example_sheet_mappings:
                rules = await self.rule_generator.generate_rules(
                    task.example_source,
                    task.example_target,
                    mapping.source_sheet,
                    mapping.target_sheet
                )
                all_rules.extend(rules)
                
                # Find corresponding task mapping
                for task_mapping in task.sheet_mappings:
                    if (task_mapping.source_sheet == mapping.source_sheet and
                        task_mapping.target_sheet == mapping.target_sheet):
                        task_mapping.rules = rules
                        break
            
            # Store all generated rules in task context
            task.context["generated_rules"] = all_rules
                
        except Exception as e:
            logger.exception(f"Rule generation failed: {str(e)}")
    
    async def _apply_rule(self, task: Task, mapping: SheetMapping, rule: Dict[str, Any]) -> bool:
        """Apply a single rule to a sheet mapping."""
        try:
            # Get rule executor from context
            executor = task.context.get("rule_executor")
            if not executor:
                logger.error("No rule executor available")
                return False
            
            # Execute rule
            return await executor.execute(rule, {
                "task": task,
                "mapping": mapping,
                "source_data": task.context.get("source_data", {}),
                "target_data": task.context.get("target_data", {})
            })
            
        except Exception as e:
            logger.exception(f"Rule application failed: {str(e)}")
            return False
    
    async def validate(self, task: Task) -> bool:
        """Validate task requirements."""
        try:
            # Basic validation is done in MigrationTask.__post_init__
            return True
        except Exception as e:
            logger.exception(f"Task validation failed: {str(e)}")
            return False