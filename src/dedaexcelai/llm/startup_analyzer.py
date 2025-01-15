from openai import OpenAI
from typing import Optional, List, Tuple, Dict, Any
import openpyxl
import re

from ..logger import get_logger, blue, green, yellow, red, magenta, cyan
from ..excel.cell_operations import get_cell_value, extract_cell_references
from ..excel.startup_rules import get_startup_days_override

logger = get_logger()

class StartupDaysAnalyzer:
    """Analyzer for determining startup days from Excel formulas."""
    
    def __init__(self, api_key: Optional[str] = None):
        """Initialize with optional OpenAI API key."""
        self.api_key = api_key
        self.client = OpenAI(api_key=api_key) if api_key else None
    
    def convert_excel_ref(self, ref: str) -> Tuple[int, int]:
        """Convert Excel reference (e.g. 'A1' or '$A$1') to row, col."""
        ref = ref.replace('$', '')  # Remove $ signs
        match = re.match(r'([A-Z]+)(\d+)', ref)
        if not match:
            raise ValueError("Invalid Excel reference: {}".format(ref))
            
        col_str, row_str = match.groups()
        
        # Convert column letters to number (A=1, B=2, etc.)
        col = 0
        for char in col_str:
            col = col * 26 + (ord(char) - ord('A') + 1)
            
        return int(row_str), col
    
    def get_surrounding_context(self, row: int, col: int, sheet: openpyxl.worksheet.worksheet.Worksheet, radius: int = 2) -> Dict[str, Any]:
        """Get values from cells surrounding the target cell."""
        context = {}
        for r in range(max(1, row - radius), min(sheet.max_row + 1, row + radius + 1)):
            for c in range(max(1, col - radius), min(sheet.max_column + 1, col + radius + 1)):
                if r != row or c != col:  # Skip the target cell itself
                    cell = sheet.cell(row=r, column=c)
                    if cell.value is not None:
                        key = f"{chr(64 + c)}{r}"  # Convert col number back to letter
                        context[key] = str(cell.value)
        return context
    
    def sheet_to_json(self, sheet: openpyxl.worksheet.worksheet.Worksheet, max_rows: int = 50) -> List[Dict[str, Any]]:
        """Convert Excel sheet to JSON format."""
        result = []
        headers = []
        
        # Get headers from first row
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            headers.append(cell.value or f"Column{col}")
            
        # Convert rows to dictionaries
        for row in range(2, min(sheet.max_row + 1, max_rows + 1)):
            row_data = {}
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                value = get_cell_value(cell)
                if value is not None:
                    row_data[headers[col-1]] = value
            if row_data:  # Only add non-empty rows
                result.append(row_data)
                
        return result
    
    def get_formula_context(self, formula: str,
                          primitive_formulas: openpyxl.worksheet.worksheet.Worksheet,
                          primitive_data: openpyxl.worksheet.worksheet.Worksheet,
                          row: int,
                          schema_sheet: openpyxl.worksheet.worksheet.Worksheet,
                          schema_formulas: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, Any]:
        """Get all relevant context for the formula."""
        # Convert sheets to JSON
        primitive_json = self.sheet_to_json(primitive_data)
        schema_json = self.sheet_to_json(schema_sheet)
        
        # Get basic formula info
        refs = extract_cell_references(formula)
        logger.debug("Found {} references", len(refs))
        
        # Build context
        context = {
            'formula': formula,
            'service_element': get_cell_value(schema_sheet.cell(row=row, column=2)),
            'primitive_sheet': primitive_json,
            'schema_sheet': schema_json,
            'references': []
        }
        
        # Add referenced cells
        for sheet_name, cell_ref in refs:
            try:
                row_num, col_num = self.convert_excel_ref(cell_ref)
                if sheet_name.upper() == 'PRIMITIVE':
                    value = get_cell_value(primitive_data.cell(row=row_num, column=col_num))
                    formula_value = primitive_formulas.cell(row=row_num, column=col_num).value
                    surrounding = self.get_surrounding_context(row_num, col_num, primitive_data)
                else:
                    value = get_cell_value(schema_sheet.cell(row=row_num, column=col_num))
                    formula_value = schema_formulas.cell(row=row_num, column=col_num).value
                    surrounding = self.get_surrounding_context(row_num, col_num, schema_sheet)
                    
                context['references'].append({
                    'cell': cell_ref,
                    'sheet': sheet_name,
                    'value': value,
                    'formula': formula_value,
                    'surrounding': surrounding
                })
                
                logger.debug("Cell {}: value={}, formula={}", 
                           cell_ref, str(value)[:50], str(formula_value)[:50])
                
            except Exception as e:
                logger.error("Error processing cell reference {}: {}", cell_ref, str(e))
        
        return context
    
    def analyze_startup_days(self, formula: str,
                           primitive_formulas: openpyxl.worksheet.worksheet.Worksheet,
                           primitive_data: openpyxl.worksheet.worksheet.Worksheet,
                           row: int,
                           schema_sheet: openpyxl.worksheet.worksheet.Worksheet,
                           schema_formulas: openpyxl.worksheet.worksheet.Worksheet,
                           filename: str,
                           element_type: str,
                           wbs_type: str) -> Optional[float]:
        """
        Analyze formula to determine startup days using OpenAI.
        """
        try:
            logger.debug(f"Analyzing formula for startup days")
            logger.debug(f"Element type: {yellow(element_type)}, WBS type: {yellow(wbs_type)}")
            
            # Early return if not a SubElement or not eligible WBS type
            if element_type != 'SubElement':
                logger.debug(f"Skipping - not a SubElement (type: {red(element_type)})")
                return None
                
            if wbs_type not in ['Fixed Optional', 'Fixed Mandatory']:
                logger.debug(f"Skipping - WBS type not eligible: {red(wbs_type)}")
                return None
            
            # Check for special cases based on filename
            logger.info(f"Checking for special cases with filename: {blue(filename)}")
            override_days = get_startup_days_override(filename, primitive_data, primitive_formulas, 
                                                    formula, element_type, wbs_type)
            if override_days is not None:
                logger.info(f"Using override value: {green(str(override_days))} days")
                return override_days
            
            if not self.client:
                logger.warning("No OpenAI client available")
                return None
                
            # Get formula context
            context = self.get_formula_context(formula, primitive_formulas, primitive_data, row, schema_sheet, schema_formulas)
            logger.debug("Got context with {} references", len(context['references']))
            
            # Build prompt with context
            refs_text = []
            for ref in context['references']:
                ref_text = [f"- {ref['sheet']} {ref['cell']}:"]
                ref_text.append(f"  Value: {ref['value']}")
                ref_text.append(f"  Formula: {ref['formula']}")
                if 'surrounding' in ref:
                    ref_text.append("  Surrounding cells:")
                    for pos, val in ref['surrounding'].items():
                        ref_text.append(f"    {pos}: {val}")
                refs_text.append("\n".join(ref_text))
            
            refs_text = "\n".join(refs_text)
            primitive_data = "\n".join(str(row) for row in context['primitive_sheet'][:10])
            schema_data = "\n".join(str(row) for row in context['schema_sheet'][:10])
            
            prompt = f"""You have access to Excel sheets in JSON format. Find the number of startup days for this service.

Service Element: {context['service_element']}
Formula being analyzed: {context['formula']}

Referenced cells and their context:
{refs_text}

PRIMITIVE sheet data (first 10 rows):
{primitive_data}

SCHEMA sheet data (first 10 rows):
{schema_data}

Question: Looking at the formula and referenced cells, what is the number of startup days?
Specifically:
1. Look for cells containing "giorni" or "pari a n giorni" in the surrounding text
2. When you find such text, look at the numeric value in the referenced cell
3. Keep the exact decimal value (like 4.44)
4. The text about days should be near the cell with the numeric value

IMPORTANT: Return ONLY a single decimal number between 0 and 365, keeping all decimal places. Just the number, no text.
Examples: '5.25' or '42.33' or '120.0'"""
            
            logger.debug("Calling OpenAI API...")
            response = self.client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are a calculator that only outputs decimal numbers. Never explain or add text to your response."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0,
                max_tokens=10
            )
            
            # Extract number from response
            try:
                days = float(response.choices[0].message.content.strip())
                logger.info(f"Determined startup days: {green(str(days))}")
                return days
            except ValueError:
                logger.error("Failed to parse response as number")
                return None
                
            return None
            
        except Exception as e:
            logger.error(f"Error analyzing startup days: {str(e)}")
            return None
