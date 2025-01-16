"""Services for Excel processing."""
from typing import Optional, Tuple, List
import openpyxl
from .models import Element, ElementType, CostType, CostMapping, FixedCostMapping, FeeCostMapping
from .cell_operations import get_cell_value
from ..logger import get_logger

logger = get_logger()

class ElementService:
    """Service for handling elements."""
    
    @staticmethod
    def create_element(row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                      primitive_sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> Optional[Element]:
        """Create an Element from a worksheet row."""
        try:
            element_type_str = get_cell_value(source_sheet.cell(row=row, column=1))
            if not element_type_str:
                return None
                
            element_type = ElementType(element_type_str)
            name = get_cell_value(source_sheet.cell(row=row, column=2))
            
            if not name or str(name).strip().startswith('-'):
                return None
                
            # Determine cost type
            cost_type = ElementService._determine_cost_type(row, source_sheet, primitive_sheet)
            
            # Calculate length for Elements
            length = None
            if element_type == ElementType.ELEMENT:
                length = len(str(name))
            
            return Element(
                name=name,
                element_type=element_type,
                cost_type=cost_type,
                row=row,
                length=length
            )
            
        except Exception as e:
            logger.error("Error creating element from row {}: {}", row, str(e))
            return None
    
    @staticmethod
    def _determine_cost_type(row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                           primitive_sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> CostType:
        """Determine the cost type for a row."""
        from .structure_analyzer import determine_cost_type
        cost_type_str = determine_cost_type(row, source_sheet, primitive_sheet)
        return CostType(cost_type_str)

class CostService:
    """Service for handling costs."""
    
    def __init__(self, current_row: int):
        self.current_row = current_row
        self.fixed_mapping = FixedCostMapping()
        self.fee_mapping = FeeCostMapping()
    
    def process_costs(self, element: Element, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                     target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Process costs for an element."""
        if element.element_type == ElementType.SUB_ELEMENT:
            self._process_subelement_costs(element, source_sheet, target_sheet)
        elif element.element_type == ElementType.ELEMENT and element.cost_type.is_fixed:
            self._process_element_costs(element, source_sheet, target_sheet)
    
    def _process_subelement_costs(self, element: Element, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                                target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Process costs for a SubElement."""
        mapping = self.fixed_mapping if element.cost_type.is_fixed else self.fee_mapping
        
        # Copy costs and prices
        self._copy_value(element.row, source_sheet, target_sheet, mapping.source_cost, mapping.target_cost)
        self._copy_value(element.row, source_sheet, target_sheet, mapping.source_price, mapping.target_price)
        
        # Set margin
        margin_cell = target_sheet.cell(row=self.current_row, column=mapping.target_margin)
        margin_cell.value = mapping.margin_value
        margin_cell.number_format = '0.00%'
        
        logger.debug("Processed costs for SubElement row {} - Type: {}", element.row, element.cost_type)
    
    def _process_element_costs(self, element: Element, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                             target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Process costs for an Element."""
        try:
            if element.cost_type.is_fixed:
                subelements = self._get_subelements_range(element.row, target_sheet)
                if subelements:
                    start_row, end_row = subelements
                    formula = f'=SUM(L{start_row}:L{end_row})'
                    target_sheet.cell(row=self.current_row, column=self.fixed_mapping.target_cost).value = formula
                    logger.debug("Set sum formula for Element row {}: {}", element.row, formula)
        except Exception as e:
            logger.error("Error processing Element costs for row {}: {}", element.row, str(e))
    
    def _get_subelements_range(self, element_row: int, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[Tuple[int, int]]:
        """Get the range of rows containing subelements for an element."""
        if not sheet:
            logger.error("Sheet is None")
            return None
            
        try:
            start_row = element_row + 1
            end_row = start_row
            
            if not hasattr(sheet, 'max_row'):
                logger.error("Sheet has no max_row attribute")
                return None
                
            for row in range(start_row, sheet.max_row + 1):
                try:
                    cell = sheet.cell(row=row, column=3)  # Object column
                    if not cell:
                        logger.error("Cell is None at row {}", row)
                        continue
                        
                    cell_type = get_cell_value(cell)
                    if cell_type == ElementType.ELEMENT.value or not cell_type:
                        break
                    if cell_type == ElementType.SUB_ELEMENT.value:
                        end_row = row
                        logger.debug("Found SubElement at row {}", row)
                except Exception as cell_error:
                    logger.error("Error accessing cell at row {}: {}", row, str(cell_error))
                    continue
            
            if end_row >= start_row:
                logger.debug("Found subelements range: {} to {}", start_row, end_row)
                return (start_row, end_row)
            logger.debug("No subelements found for row {}", element_row)
            return None
            
        except Exception as e:
            logger.error("Error getting subelements range: {}", str(e))
            return None
    
    def _copy_value(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                   target_sheet: openpyxl.worksheet.worksheet.Worksheet, source_col: int, target_col: int) -> None:
        """Copy value from source to target cell."""
        value = get_cell_value(source_sheet.cell(row=row, column=source_col))
        if value is not None:
            target_cell = target_sheet.cell(row=self.current_row, column=target_col)
            target_cell.value = value
            if isinstance(value, (int, float)):
                target_cell.number_format = '#,##0.0000'
