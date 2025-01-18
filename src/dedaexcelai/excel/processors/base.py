"""Excel sheet processors."""
from abc import ABC, abstractmethod
import openpyxl
from typing import Optional, List
from dedaexcelai.logger import get_logger
from dedaexcelai.excel.services.element_service import ElementService
from dedaexcelai.excel.services.cost_service import CostService, create_cost_processor
from dedaexcelai.excel.models.elements import ElementType, Element
from dedaexcelai.excel.utils.styling import set_euro_format

logger = get_logger()

class SheetProcessor(ABC):
    """Base class for sheet processors."""
    
    @abstractmethod
    def process(self, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
               target_sheet: openpyxl.worksheet.worksheet.Worksheet,
               source_formulas: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> bool:
        """Process source sheet and write to target sheet."""
        pass

class PrimitiveSheetProcessor(SheetProcessor):
    """Processor for PRIMITIVE sheet."""
    
    def process(self, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
               target_sheet: openpyxl.worksheet.worksheet.Worksheet,
               source_formulas: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> bool:
        """Copy PRIMITIVE sheet as-is."""
        try:
            logger.info("Processing PRIMITIVE sheet...")
            for row in source_sheet.values:
                target_sheet.append(row)
            return True
        except Exception as e:
            logger.error("Failed to process PRIMITIVE sheet: {}", str(e))
            return False

class SchemaSheetProcessor(SheetProcessor):
    """Processor for SCHEMA sheet."""
    
    def __init__(self, startup_analyzer=None, filename: str = ""):
        self.startup_analyzer = startup_analyzer
        self.filename = filename
        self.current_output_row = 2  # Start after headers
        self.primitive_data = None
        self.primitive_formulas = None
    
    def setup_headers(self, target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Set up headers with green background."""
        headers = [
            'Lenght', 'Product Element', 'Object', 'Type', 'GG Startup', 'RU', 'RU Qty',
            'RU Unit of measure', 'Q.ty min', 'Q.ty MAX', '%Sconto MAX',
            'Startup Costo', 'Startup Margine', 'Startup Prezzo',
            'Canone Costo Mese', 'Canone Margine', 'Canone Prezzo Mese',
            'Extended Description', 'Profit Center Prevalente', 'Status', 'Note'
        ]
        
        green_fill = openpyxl.styles.PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = target_sheet.cell(row=1, column=col, value=header)
            cell.fill = green_fill
            cell.font = openpyxl.styles.Font(bold=True)
    
    def process_row(self, row: int, source_sheet: Optional[openpyxl.worksheet.worksheet.Worksheet],
                   target_sheet: Optional[openpyxl.worksheet.worksheet.Worksheet],
                   source_formulas: Optional[openpyxl.worksheet.worksheet.Worksheet]) -> None:
        """Process a single row."""
        if not source_sheet or not target_sheet:
            logger.error("Source or target sheet is None")
            return
            
        # Create element from row with primitive data for formula evaluation
        element = ElementService.create_element(row, source_sheet, self.primitive_data)
        if not element:
            return
            
        try:
            logger.debug("Processing {} row {} - Type: {}", element.element_type.value, row, element.cost_type.value)
        
            # Set basic properties
            if element.length is not None:
                target_sheet.cell(row=self.current_output_row, column=1).value = element.length
                
            # Set element name and formatting
            element_cell = target_sheet.cell(row=self.current_output_row, column=2)
            element_cell.value = element.name
            element_cell.font = openpyxl.styles.Font(
                bold=(element.element_type == ElementType.ELEMENT),
                italic=(element.element_type == ElementType.SUB_ELEMENT)
            )
            
            # Set element and cost types
            target_sheet.cell(row=self.current_output_row, column=3).value = element.element_type.value
            target_sheet.cell(row=self.current_output_row, column=4).value = element.cost_type.value
            
            # Process costs
            cost_service = CostService(self.current_output_row)
            cost_service.process_costs(element, source_sheet, target_sheet)
            
            # Process startup days for Fixed costs
            if element.cost_type.is_fixed and source_formulas:
                self._process_startup_days(element, source_sheet, target_sheet, source_formulas)
            
            self.current_output_row += 1
            
        except Exception as e:
            logger.error("Error processing row {}: {}", row, str(e))
    
    def _process_startup_days(self, element: Element, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                            target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                            source_formulas: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Process startup days for Fixed costs."""
        if not self.startup_analyzer:
            logger.warning("No startup analyzer available")
            return
            
        formula = source_formulas.cell(row=element.row, column=8).value  # Column H
        if not formula:
            return
            
        # Convert float to string if needed
        if isinstance(formula, float):
            formula = str(formula)
        elif not isinstance(formula, str):
            logger.warning("Formula is not a string or float: {}", type(formula))
            return
            
        try:
            if not self.primitive_data or not self.primitive_formulas:
                logger.warning("Missing primitive sheets")
                return
                
            startup_days = self.startup_analyzer.analyze_startup_days(
                formula,
                self.primitive_formulas,
                self.primitive_data,
                element.row,
                source_sheet,
                source_formulas,
                self.filename,
                element.element_type.value,
                element.cost_type.value
            )
            
            if startup_days is not None:
                logger.info("Got startup days: {}", startup_days)
                target_sheet.cell(row=self.current_output_row, column=5).value = startup_days
                element.startup_days = startup_days
            else:
                logger.warning("No startup days returned for row {}", element.row)
                
        except Exception as e:
            logger.error("Error analyzing startup days for row {}: {}", element.row, str(e))
    
    def process(self, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
               target_sheet: openpyxl.worksheet.worksheet.Worksheet,
               source_formulas: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> bool:
        """Process SCHEMA sheet with new format."""
        try:
            logger.info("Processing SCHEMA sheet...")
            
            # Set up headers
            self.setup_headers(target_sheet)
            
            # Find header row in source
            from dedaexcelai.excel.utils.structure import find_header_row
            header_row = find_header_row(source_sheet)
            
            # Process rows
            for row in range(header_row + 1, source_sheet.max_row + 1):
                self.process_row(row, source_sheet, target_sheet, source_formulas)
                
            return True
            
        except Exception as e:
            logger.error("Failed to process SCHEMA sheet: {}", str(e))
            return False
