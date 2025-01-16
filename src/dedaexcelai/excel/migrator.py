"""Excel migration module."""
from typing import Optional
import openpyxl
from dedaexcelai.logger import get_logger
from dedaexcelai.excel.core.factory import create_sheet_processor
from dedaexcelai.excel.core.workbook import load_workbooks, create_output_workbook
from dedaexcelai.llm.startup_analyzer import StartupDaysAnalyzer

logger = get_logger()

def migrate_excel(input_path: str, output_path: str, openai_key: Optional[str] = None) -> bool:
    """Migrate Excel file to new format."""
    try:
        # Initialize startup analyzer if key provided
        startup_analyzer = None
        if openai_key:
            logger.info("Initializing StartupDaysAnalyzer...")
            startup_analyzer = StartupDaysAnalyzer(openai_key)
            logger.info("Successfully initialized StartupDaysAnalyzer with GPT-4")
        
        # Load workbooks
        logger.info("Loading input workbook...")
        source_workbook = load_workbooks(input_path)
        if not source_workbook:
            return False
            
        # Create output workbook
        logger.info("Creating new workbook...")
        target_workbook = create_output_workbook()
        
        # Process PRIMITIVE sheet
        logger.info("Creating PRIMITIVE processor...")
        primitive_processor = create_sheet_processor("PRIMITIVE")
        if not primitive_processor:
            logger.error("Failed to create PRIMITIVE processor")
            return False
            
        primitive_sheet = source_workbook["PRIMITIVE"]
        target_primitive = target_workbook.create_sheet("PRIMITIVE")
        if not primitive_processor.process(primitive_sheet, target_primitive):
            logger.error("Failed to process PRIMITIVE sheet")
            return False
            
        # Store PRIMITIVE sheets for startup analyzer
        primitive_data = primitive_sheet
        primitive_formulas = source_workbook["PRIMITIVE_FORMULAS"] if "PRIMITIVE_FORMULAS" in source_workbook else None
        
        # Process SCHEMA sheet
        logger.info("Creating SCHEMA processor with startup analyzer...")
        schema_processor = create_sheet_processor("SCHEMA", startup_analyzer, input_path)
        if not schema_processor:
            logger.error("Failed to create SCHEMA processor")
            return False
            
        # Set primitive sheets for startup analyzer
        if hasattr(schema_processor, 'primitive_data'):
            schema_processor.primitive_data = primitive_data
        if hasattr(schema_processor, 'primitive_formulas'):
            schema_processor.primitive_formulas = primitive_formulas
            
        logger.info("Processing SCHEMA sheet with startup analyzer...")
        schema_sheet = source_workbook["SCHEMA"]
        target_schema = target_workbook.create_sheet("SCHEMA")
        schema_formulas = source_workbook["SCHEMA_FORMULAS"] if "SCHEMA_FORMULAS" in source_workbook else None
        
        if not schema_processor.process(schema_sheet, target_schema, schema_formulas):
            logger.error("Failed to process SCHEMA sheet")
            return False
            
        # Remove default sheet
        if "Sheet" in target_workbook.sheetnames:
            target_workbook.remove(target_workbook["Sheet"])
            
        # Save output workbook
        target_workbook.save(output_path)
        logger.info("Successfully saved output workbook to {}", output_path)
        
        return True
        
    except Exception as e:
        logger.error("Failed to migrate Excel file: {}", str(e))
        return False
