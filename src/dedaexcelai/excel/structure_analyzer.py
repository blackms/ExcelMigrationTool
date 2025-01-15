from typing import Tuple
import openpyxl
from .cell_operations import get_cell_value, is_empty_or_dashes
from ..logger import get_logger, blue, yellow, magenta, green

logger = get_logger()

def find_element_catalog_interval(sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int) -> Tuple[int, int]:
    """
    Find the interval for the current element catalog by looking for separator rows
    Returns: (start_row, end_row)
    """
    # First, go back to find the start of this catalog if we're in the middle
    while start_row > 1:
        service_element = get_cell_value(sheet.cell(row=start_row - 1, column=2))
        if is_empty_or_dashes(service_element) and isinstance(service_element, str) and "-" in service_element:
            break
        start_row -= 1
    
    # Then find the end (next separator row)
    end_row = start_row
    max_row = sheet.max_row
    
    while end_row <= max_row:
        service_element = get_cell_value(sheet.cell(row=end_row, column=2))
        if end_row > start_row and is_empty_or_dashes(service_element) and isinstance(service_element, str) and "-" in service_element:
            end_row -= 1  # Don't include the separator row
            break
        end_row += 1
    
    if end_row > max_row:
        end_row = max_row
    
    return (start_row, end_row)

def determine_cost_type(row: int, sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
    """
    Determine the cost type based on WBS column and mandatory flag within the element catalog interval
    Returns: 
        - "Fee Optional" for CANONE (non-mandatory)
        - "Fee Mandatory" for CANONE (mandatory)
        - "Fixed Optional" for FIXED (non-mandatory)
        - "Fixed Mandatory" for FIXED (mandatory)
    """
    # Find the interval for the current element catalog
    interval_start, interval_end = find_element_catalog_interval(sheet, row)
    
    # Get the service element name for better logging
    service_element = get_cell_value(sheet.cell(row=interval_start, column=2))
    logger.debug(f"Element catalog '{yellow(service_element)}' (rows {blue(str(interval_start))}-{blue(str(interval_end))})")
    
    # Look for FIXED or CANONE in the WBS column (column D) and mandatory flag in column F
    wbs_type = None
    is_mandatory = False
    
    # First check if the current row is mandatory (column F)
    mandatory_value = get_cell_value(sheet.cell(row=row, column=6))  # Column F is 6
    if isinstance(mandatory_value, str) and mandatory_value.upper().strip() == "M":
        is_mandatory = True
        logger.debug(f"Found Mandatory flag at row {blue(str(row))}")
    
    # Then check the WBS type within the interval
    for r in range(interval_start, interval_end + 1):
        wbs_value = get_cell_value(sheet.cell(row=r, column=4))  # Column D is 4
        if isinstance(wbs_value, str):
            wbs_value = wbs_value.upper().strip()
            if wbs_value == "FIXED":
                wbs_type = "FIXED"
                logger.debug(f"Found {magenta('FIXED')} type at row {blue(str(r))}")
                break
            elif wbs_value == "CANONE":
                wbs_type = "CANONE"
                logger.debug(f"Found {magenta('CANONE')} type at row {blue(str(r))}")
                break
    
    # Determine final cost type based on WBS type and mandatory flag
    if wbs_type == "FIXED":
        cost_type = "Fixed Mandatory" if is_mandatory else "Fixed Optional"
        logger.debug(f"Using {green(cost_type)} based on WBS type and mandatory flag")
        return cost_type
    elif wbs_type == "CANONE":
        cost_type = "Fee Mandatory" if is_mandatory else "Fee Optional"
        logger.debug(f"Using {green(cost_type)} based on WBS type and mandatory flag")
        return cost_type
    else:
        default_type = "Fee Mandatory" if is_mandatory else "Fee Optional"
        logger.warning(f"No WBS type found in interval, defaulting to {yellow(default_type)}")
        return default_type

def find_header_row(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Find the row containing headers"""
    for row in range(1, sheet.max_row + 1):
        if get_cell_value(sheet.cell(row=row, column=1)) == "Portfolio":
            return row
    return 1
