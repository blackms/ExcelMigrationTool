import openpyxl
from typing import Optional
from pathlib import Path
import tempfile
import shutil
import os

from ..logger import get_logger

logger = get_logger()

class WorkbookHandler:
    """Handles workbook operations like loading, saving, and sheet management."""
    
    def __init__(self, input_file: str, output_file: str, template_file: str = None):
        """Initialize with input and output file paths."""
        self.input_file = input_file
        self.output_file = output_file
        self.template_file = template_file
        self._filename = os.path.basename(input_file)
        self.input_wb = None
        self.input_wb_formulas = None  # Workbook with formulas
        self.output_wb = None
        
    @property
    def filename(self) -> str:
        """Get the input filename without path."""
        return self._filename
        
    def load_workbooks(self) -> bool:
        """Load input and create output workbooks."""
        try:
            logger.info("Loading input workbook...")
            self.input_wb = openpyxl.load_workbook(self.input_file, data_only=True)
            self.input_wb_formulas = openpyxl.load_workbook(self.input_file, data_only=False)
            return True
        except Exception as e:
            logger.error(f"Failed to load workbooks: {str(e)}")
            return False
            
    def create_output_workbook(self) -> bool:
        """Create new output workbook."""
        try:
            logger.info("Creating new workbook...")
            self.output_wb = openpyxl.Workbook()
            if 'Sheet' in self.output_wb.sheetnames:
                del self.output_wb['Sheet']
            return True
        except Exception as e:
            logger.error(f"Failed to create output workbook: {str(e)}")
            return False
            
    def save_workbook(self) -> bool:
        """Save the output workbook."""
        try:
            if not self.output_wb:
                logger.error("No output workbook to save")
                return False
                
            logger.info("Saving output workbook...")
            self.output_wb.save(self.output_file)
            logger.info(f"Successfully saved to: {self.output_file}")
            return True
        except Exception as e:
            logger.error(f"Failed to save workbook: {str(e)}")
            return False
            
    def get_sheet(self, sheet_name: str, data_only: bool = True) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """Get sheet from input workbook."""
        try:
            if data_only:
                return self.input_wb[sheet_name] if sheet_name in self.input_wb.sheetnames else None
            else:
                return self.input_wb_formulas[sheet_name] if sheet_name in self.input_wb_formulas.sheetnames else None
        except Exception as e:
            logger.error(f"Failed to get sheet {sheet_name}: {str(e)}")
            return None
        
    def create_sheet(self, sheet_name: str, index: int = None) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """Create a new sheet in the output workbook."""
        try:
            if not self.output_wb:
                self.output_wb = openpyxl.Workbook()
                
            if sheet_name in self.output_wb.sheetnames:
                return self.output_wb[sheet_name]
                
            if index is not None:
                return self.output_wb.create_sheet(sheet_name, index)
            else:
                return self.output_wb.create_sheet(sheet_name)
        except Exception as e:
            logger.error(f"Failed to create sheet {sheet_name}: {str(e)}")
            return None
