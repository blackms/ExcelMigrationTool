import openpyxl
from typing import Optional
from dedaexcelai.excel.utils.cell import get_cell_value
from dedaexcelai.logger import get_logger

logger = get_logger()

def find_header_row(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Find the header row in a sheet."""
    for row in range(1, min(10, sheet.max_row + 1)):  # Only check first 10 rows
        cell = sheet.cell(row=row, column=1)
        if get_cell_value(cell) == 'Type':
            return row
    return 1  # Default to first row if not found

def column_letter_to_number(column_letter: str) -> int:
    """Convert Excel column letter to number (1-based)."""
    result = 0
    for char in column_letter:
        result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def evaluate_sum_formula(formula: str, sheet: openpyxl.worksheet.worksheet.Worksheet) -> float:
    """Evaluate a SUM formula like =SUM(H5:H10)."""
    try:
        if not formula.upper().startswith('=SUM('):
            return 0
            
        # Extract range from SUM(X1:Y2)
        range_part = formula[5:-1]  # Remove =SUM( and )
        start_ref, end_ref = range_part.split(':')
        
        # Get start and end coordinates
        start_col = ''.join(c for c in start_ref if c.isalpha())
        start_row = int(''.join(c for c in start_ref if c.isdigit()))
        end_col = ''.join(c for c in end_ref if c.isalpha())
        end_row = int(''.join(c for c in end_ref if c.isdigit()))
        
        start_col_num = column_letter_to_number(start_col)
        end_col_num = column_letter_to_number(end_col)
        
        # Sum all cells in range
        total = 0
        for row in range(start_row, end_row + 1):
            for col in range(start_col_num, end_col_num + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)):
                    total += float(cell_value)
                    
        logger.debug(f"SUM formula result: {total}")
        return total
        
    except Exception as e:
        logger.error(f"Error evaluating SUM formula: {str(e)}")
        return 0

def get_cell_value_by_ref(cell_ref: str, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Optional[float]:
    """Get numeric value from a cell reference like 'A1'."""
    try:
        # Split reference into column letters and row number
        col_letters = ''.join(c for c in cell_ref if c.isalpha())
        row_num = ''.join(c for c in cell_ref if c.isdigit())
        
        if not col_letters or not row_num:
            logger.error(f"Invalid cell reference: {cell_ref}")
            return None
            
        col = column_letter_to_number(col_letters)
        row = int(row_num)
        
        logger.debug(f"Accessing cell - Col: {col_letters}({col}), Row: {row}")
        cell_value = sheet.cell(row=row, column=col).value
        
        if cell_value is None:
            logger.error(f"Cell value is None at {col_letters}{row}")
            return None
            
        # If it's a number, return it
        if isinstance(cell_value, (int, float)):
            return float(cell_value)
            
        # If it's a formula, evaluate it
        if isinstance(cell_value, str) and cell_value.startswith('='):
            logger.debug(f"Found nested formula: {cell_value}")
            # Handle division formula (e.g., =S25/B13)
            if '/' in cell_value:
                parts = cell_value[1:].split('/')  # Remove = and split
                numerator = get_cell_value_by_ref(parts[0].strip().replace('$', ''), sheet)
                denominator = get_cell_value_by_ref(parts[1].strip().replace('$', ''), sheet)
                if numerator is not None and denominator is not None and denominator != 0:
                    return numerator / denominator
            
        return None
    except Exception as e:
        logger.error(f"Error getting cell value: {str(e)}")
        return None

def evaluate_primitive_formula(formula: str, primitive_sheet: openpyxl.worksheet.worksheet.Worksheet) -> float:
    """Evaluate a formula that references the PRIMITIVE sheet."""
    try:
        if not formula or not isinstance(formula, str):
            return 0
            
        # Handle PRIMITIVE!X##*PRIMITIVE!Y## format
        if 'PRIMITIVE!' in formula:
            parts = formula.split('*')
            result = 1
            for part in parts:
                if 'PRIMITIVE!' in part:
                    cell_ref = part.strip().replace('PRIMITIVE!', '')
                    value = get_cell_value_by_ref(cell_ref, primitive_sheet)
                    if value is not None:
                        result *= value
                    else:
                        return 0
                        
            logger.debug(f"Formula result: {result}")
            return result
            
        return 0
    except Exception as e:
        logger.error(f"Error evaluating primitive formula: {str(e)}")
        return 0

def determine_cost_type(row: int, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                       primitive_sheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> str:
    """Determine the cost type based on column H and I values."""
    # Check if this is a STARTUP subelement
    element_name = get_cell_value(sheet.cell(row=row, column=2))  # Column B
    element_type = get_cell_value(sheet.cell(row=row, column=3))  # Column C
    is_startup_subelement = (element_type == "SubElement" and 
                           isinstance(element_name, str) and 
                           "STARTUP" in element_name.upper())
    
    h_cell = sheet.cell(row=row, column=8)  # Column H
    i_cell = sheet.cell(row=row, column=9)  # Column I
    
    # Get formulas
    h_formula = h_cell.value
    i_formula = i_cell.value
    
    # Evaluate formulas
    h_value = h_formula
    i_value = i_formula
    
    if isinstance(h_formula, str):
        if 'PRIMITIVE!' in h_formula:
            h_value = evaluate_primitive_formula(h_formula, primitive_sheet) if primitive_sheet else 0
        elif h_formula.upper().startswith('=SUM('):
            h_value = evaluate_sum_formula(h_formula, sheet)
            
    if isinstance(i_formula, str):
        if 'PRIMITIVE!' in i_formula:
            i_value = evaluate_primitive_formula(i_formula, primitive_sheet) if primitive_sheet else 0
        elif i_formula.upper().startswith('=SUM('):
            i_value = evaluate_sum_formula(i_formula, sheet)
    
    # Convert final values to float
    try:
        h_numeric = float(h_value if h_value is not None else 0)
    except (ValueError, TypeError):
        h_numeric = 0
        
    try:
        i_numeric = float(i_value if i_value is not None else 0)
    except (ValueError, TypeError):
        i_numeric = 0
    
    # Log both formula and calculated values
    logger.debug(f"Row {row} - Formula - H: {h_formula}, I: {i_formula}")
    logger.debug(f"Row {row} - Calculated - H: {h_value}, I: {i_value}")
    logger.debug(f"Row {row} - Numeric - H: {h_numeric}, I: {i_numeric}")
    
    if h_numeric != 0 or is_startup_subelement:
        return 'Fixed Optional'
    elif i_numeric != 0:
        return 'Fee Optional'
    # Default to Fee Optional if no values found
    return 'Fee Optional'
