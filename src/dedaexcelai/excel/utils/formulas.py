from typing import Dict, Optional, Tuple, List
import openpyxl
from dedaexcelai.logger import get_logger

logger = get_logger()

class ColumnFormulas:
    """Handles column formulas and mappings for Excel processing."""
    
    @staticmethod
    def get_subelements_range(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                            element_row: int) -> Optional[Tuple[int, int]]:
        """Get the range of rows containing subelements for an element."""
        try:
            element_type = sheet.cell(row=element_row, column=3).value  # Object column
            if element_type != 'Element':
                logger.debug("Row {} is not an Element (type: {})", element_row, element_type)
                return None
                
            start_row = element_row + 1
            end_row = start_row
            
            # Find last subelement row
            for row in range(start_row, sheet.max_row + 1):
                cell_type = sheet.cell(row=row, column=3).value  # Object column
                logger.debug("Checking row {} - Type: {}", row, cell_type)
                if cell_type == 'Element' or not cell_type:
                    break
                if cell_type == 'SubElement':
                    end_row = row
                    logger.debug("Found SubElement at row {}", row)
                    
            if end_row >= start_row:
                logger.debug("Found subelements range: {} to {}", start_row, end_row)
                return (start_row, end_row)
            logger.debug("No subelements found for row {}", element_row)
            return None
            
        except Exception as e:
            logger.error("Error getting subelements range: {}", str(e))
            return None
    
    @staticmethod
    def get_element_formulas(row: int, subelements_range: Optional[Tuple[int, int]]) -> Dict[str, str]:
        """Get formulas for Element rows."""
        formulas = {}
        
        if subelements_range:
            start_row, end_row = subelements_range
            # Sum of subelements' column L values (Unatantu)
            formula = f'=SUM(L{start_row}:L{end_row})'
            formulas['L'] = formula
            logger.debug("Generated formula for row {}: {}", row, formula)
        
        return formulas
    
    @staticmethod
    def apply_formulas(target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                      target_row: int,
                      element_type: str,
                      cost_type: str,
                      source_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Apply appropriate formulas based on element and cost type."""
        try:
            logger.debug("Processing target row {} - Type: {}, Cost: {}", target_row, element_type, cost_type)
            
            if element_type == 'Element' and 'Fixed' in str(cost_type):
                # Get subelements range for sum formulas
                subelements_range = ColumnFormulas.get_subelements_range(target_sheet, target_row)
                logger.debug("Subelements range for row {}: {}", target_row, subelements_range)
                
                # Apply formulas for Fixed Elements
                formulas = ColumnFormulas.get_element_formulas(target_row, subelements_range)
                if formulas:
                    logger.debug("Generated formulas for row {}", target_row)
                    for col, formula in formulas.items():
                        cell = target_sheet.cell(row=target_row, column=ord(col)-64)
                        cell.value = formula
                        logger.info("Set {}{} = {}", col, target_row, formula)
                        
                        # Verify formula was set
                        actual_value = target_sheet.cell(row=target_row, column=ord(col)-64).value
                        logger.debug("Verified {}{} value: {}", col, target_row, actual_value)
                else:
                    logger.debug("No formulas generated for row {}", target_row)
                        
        except Exception as e:
            logger.error("Error applying formulas: {}", str(e))
