from typing import Optional, Dict, Any
import openpyxl
from ...logger import get_logger

logger = get_logger()

class FileRules:
    """Rules manager for specific file processing."""
    
    @staticmethod
    def get_column_mapping(filename: str) -> Dict[str, int]:
        """Get column mapping based on file type."""
        # Colonne generiche sempre presenti
        base_mapping = {
            "startup_cost": 11,    # Column K for Startup Costo
            "startup_margin": 12,  # Column L for Startup Margin
            "startup_price": 13,   # Column M for Startup Prezzo
        }
        
        # Colonne specifiche per COaaS_Schema
        if "COaaS_Schema" in filename:
            base_mapping.update({
                "ru": 6,          # Column F for Resource Unit
                "ru_qty": 7,      # Column G for RU Qty
                "ru_uom": 8,      # Column H for RU Unit of measure
                "qty_min": 9,     # Column I for Qty min
                "startup_days": 5, # Column E for GG Startup
                "profit_center": 18  # Column R for Profit Center
            })
        return base_mapping
    
    @staticmethod
    def get_resource_unit_rule(filename: str, element_type: str, row_data: Dict[str, Any]) -> Optional[str]:
        """Get Resource Unit value based on file-specific rules."""
        try:
            # Rules for COaaS Schema
            if "COaaS_Schema" in filename:
                if element_type == "Element":
                    logger.debug(f"Setting 'Per ogni contratto' for Element")
                    return "Per Ogni Contratto"
                elif element_type == "SubElement":
                    # Per i SubElement, ritorniamo None per usare il valore originale
                    logger.debug(f"Using original RU value for SubElement")
                    return None
            
            # Default: return None to use original value
            return None
            
        except Exception as e:
            logger.error(f"Error in resource unit rule for {filename}: {str(e)}")
            return None
    
    @staticmethod
    def get_startup_days_rule(filename: str) -> Dict[str, Any]:
        """Get startup days calculation rules for specific files."""
        rules = {
            "enabled": False,
            "source_cells": [],
            "calculation": None
        }
        
        if "COaaS_Schema" in filename:
            rules.update({
                "enabled": True,
                "source_cells": [
                    {"sheet": "PRIMITIVE", "cell": "S25"},
                    {"sheet": "PRIMITIVE", "cell": "B13"}
                ],
                "calculation": lambda s25, b13: float(s25) / float(b13) if float(b13) != 0 else None
            })
            
        return rules
    
    @staticmethod
    def apply_column_rules(filename: str, element_type: str, column: str, value: Any) -> Optional[Any]:
        """Apply rules for specific columns - File specific rules."""
        try:
            if "COaaS_Schema" in filename:
                if element_type == "Element":
                    if column == "ru":
                        return "Per ogni contratto"
                    elif column == "ru_qty":
                        return 1
                    elif column == "ru_uom":
                        return "Item"
                    elif column == "qty_min":
                        return 1
            return value
            
        except Exception as e:
            logger.error(f"Error applying column rules: {str(e)}")
            return value
    
    @staticmethod
    def get_formula_range(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                         current_row: int, 
                         column: int) -> Optional[tuple[int, int]]:
        """Get the range of SubElements for formula calculation."""
        try:
            start_row = current_row + 1
            current_indent = sheet.cell(row=current_row, column=2).alignment.indent
            
            # Cerca la fine dei SubElement
            end_row = start_row
            while end_row <= sheet.max_row:
                cell = sheet.cell(row=end_row, column=2)
                if cell.alignment.indent <= current_indent:
                    break
                end_row += 1
                
            return (start_row, end_row - 1) if end_row > start_row else None
            
        except Exception as e:
            logger.error(f"Error getting formula range: {str(e)}")
            return None
    
    @staticmethod
    def get_startup_formulas(element_type: str, 
                           cost_type: str,
                           current_row: int,
                           range_info: Optional[tuple[int, int]]) -> Dict[str, Optional[str]]:
        """Get formulas for startup columns - Generic rules for all files."""
        formulas = {
            "startup_cost": None,
            "startup_margin": None,
            "startup_price": None
        }
        
        # Regole generiche per tutti i file
        if element_type == "Element" and "Fixed" in cost_type:
            if range_info:
                start_row, end_row = range_info
                formulas.update({
                    "startup_cost": f"=SUM(K{start_row}:K{end_row})",
                    "startup_margin": f"=IFERROR(1-(K{current_row}/M{current_row}),0)",
                    "startup_price": f"=SUM(M{start_row}:M{end_row})"
                })
                
        return formulas
