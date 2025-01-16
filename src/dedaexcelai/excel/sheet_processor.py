from abc import ABC, abstractmethod
import openpyxl
from typing import Optional, Dict, Any
from ..logger import get_logger
from .cell_operations import get_cell_value, is_empty_or_dashes
from .style_manager import set_euro_format
from .column_formulas import ColumnFormulas
from .cost_processor import create_cost_processor

logger = get_logger()

class SheetProcessor(ABC):
    """Base class for sheet processors."""
    
    @abstractmethod
    def process(self, source_sheet: openpyxl.worksheet.worksheet.Worksheet, target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """Process source sheet and write to target sheet."""
        pass

class PrimitiveSheetProcessor(SheetProcessor):
    """Processor for PRIMITIVE sheet."""
    
    def process(self, source_sheet: openpyxl.worksheet.worksheet.Worksheet, target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
        """Copy PRIMITIVE sheet as-is."""
        try:
            logger.info("Processing PRIMITIVE sheet...")
            for row in source_sheet.values:
                target_sheet.append(row)
            return True
        except Exception as e:
            logger.error("Failed to process PRIMITIVE sheet: {}", str(e))
            return False

class SchemaSheetProcessor(SheetProcessor):
    """Processor for SCHEMA sheet."""
    
    def __init__(self, startup_analyzer=None, filename: str = ""):
        self.startup_analyzer = startup_analyzer
        self.filename = filename
        self.header_row = None
        self.current_output_row = 2  # Start after headers
        self.primitive_data = None
        self.primitive_formulas = None
        
    def setup_headers(self, target_sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Set up headers with green background."""
        headers = [
            'Lenght', 'Product Element', 'Object', 'Type', 'GG Startup', 'RU', 'RU Qty',
            'RU Unit of measure', 'Q.ty min', 'Q.ty MAX', '%Sconto MAX',
            'Startup Costo', 'Startup Margine', 'Startup Prezzo',
            'Canone Costo Mese', 'Canone Margine', 'Canone Prezzo Mese',
            'Extended Description', 'Profit Center Prevalente', 'Status', 'Note'
        ]
        
        green_fill = openpyxl.styles.PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = target_sheet.cell(row=1, column=col, value=header)
            cell.fill = green_fill
            cell.font = openpyxl.styles.Font(bold=True)
    
    def determine_type(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
        """Determine if cost is Fee or Fixed based on values."""
        # Check Unatantu (H) and Ricorrente (I)
        unatantu = get_cell_value(source_sheet.cell(row=row, column=8))
        ricorrente = get_cell_value(source_sheet.cell(row=row, column=9))
        
        # If Unatantu has a value and Ricorrente is 0 or empty, it's Fixed
        if (unatantu is not None and str(unatantu).strip() and 
            (ricorrente is None or str(ricorrente).strip() == '0' or str(ricorrente).strip() == '-')):
            return 'Fixed Optional'
            
        # Default to Fee Optional
        return 'Fee Optional'
    
    def process_row(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet, 
                   target_sheet: openpyxl.worksheet.worksheet.Worksheet, source_formulas: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Process a single row from source to target."""
        # Get type and service element
        element_type = get_cell_value(source_sheet.cell(row=row, column=1))  # Type column
        service_element = get_cell_value(source_sheet.cell(row=row, column=2))  # Service Element column
        
        if service_element is None or is_empty_or_dashes(service_element):
            return
            
        # Set Lenght for Elements
        if element_type == 'Element':
            text_length = len(str(service_element)) if service_element else 0
            target_sheet.cell(row=self.current_output_row, column=1).value = text_length
        
        # Copy Product Element and set formatting
        element_cell = target_sheet.cell(row=self.current_output_row, column=2)
        element_cell.value = service_element
        element_cell.font = openpyxl.styles.Font(bold=(element_type == 'Element'), italic=(element_type == 'SubElement'))
        
        # Set Object (Element/SubElement)
        target_sheet.cell(row=self.current_output_row, column=3).value = element_type
        
        # Set Type (Fee/Fixed)
        cost_type = self.determine_type(row, source_sheet)
        target_sheet.cell(row=self.current_output_row, column=4).value = cost_type
        
        # Process other columns based on mapping
        self.process_mapped_columns(row, source_sheet, target_sheet, element_type, cost_type)
        
        # Handle GG Startup for Fixed costs
        if cost_type.startswith('Fixed'):
            logger.info("Processing startup days for Fixed cost in row {}", row)
            # Get formula from Unatantu column (H)
            formula = source_formulas.cell(row=row, column=8).value if source_formulas else None
            if formula is None:
                # If no formula, try to get direct value
                formula = str(get_cell_value(source_sheet.cell(row=row, column=8)) or '')
            logger.info("Found value in Unatantu column: {}", formula)
            self.process_startup_days(row, source_sheet, target_sheet, source_formulas, formula)
        else:
            logger.debug("Skipping startup days for non-Fixed cost in row {}", row)
        
        # Apply column formulas
        logger.debug("Applying formulas for row {} - Element: {}, Cost: {}", self.current_output_row, element_type, cost_type)
        ColumnFormulas.apply_formulas(
            target_sheet=target_sheet,
            target_row=self.current_output_row,
            element_type=element_type,
            cost_type=cost_type,
            source_sheet=source_sheet
        )
        
        self.current_output_row += 1
    
    def process_mapped_columns(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet, 
                             target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                             element_type: str, cost_type: str) -> None:
        """Process columns based on mapping."""
        from .file_rules import FileRules
        
        # Get RU value based on rules
        ru_value = FileRules.get_resource_unit_rule(self.filename, element_type, {})
        if ru_value:
            target_sheet.cell(row=self.current_output_row, column=6).value = ru_value
        else:
            # Copy original Resource Unit if no rule applies
            self.copy_value(row, source_sheet, target_sheet, 3, 6)  # Resource Unit -> RU
        
        # Copy Profit Center
        self.copy_value(row, source_sheet, target_sheet, 6, 18)  # Profit Center -> Profit Center Prevalente
        
        # Process cost-specific columns
        cost_processor = create_cost_processor(cost_type, self.current_output_row)
        cost_processor.process_columns(row, source_sheet, target_sheet, element_type)
    
    def copy_value(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                  target_sheet: openpyxl.worksheet.worksheet.Worksheet, source_col: int, target_col: int) -> None:
        """Copy value from source to target cell."""
        value = get_cell_value(source_sheet.cell(row=row, column=source_col))
        if value is not None:
            target_cell = target_sheet.cell(row=self.current_output_row, column=target_col)
            target_cell.value = value
            # Set number format with 4 decimal places for numeric values
            if isinstance(value, (int, float)):
                target_cell.number_format = '#,##0.0000'
    
    def process_startup_days(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                           target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                           source_formulas: openpyxl.worksheet.worksheet.Worksheet,
                           formula: str) -> None:
        """Process startup days for a row."""
        logger.info("Processing startup days for row {}", row)
        
        if not self.startup_analyzer:
            logger.warning("No startup analyzer available")
            return
            
        if not formula:
            logger.warning("No formula provided")
            return
            
        try:
            logger.info("Calling startup analyzer...")
            # Get element type from column 1
            element_type = get_cell_value(source_sheet.cell(row=row, column=1))
            logger.debug("Processing element type: {}", element_type)
            
            if not self.primitive_data or not self.primitive_formulas:
                logger.warning("Missing primitive sheets")
                return
                
            # Get WBS type from column D
            wbs_type = get_cell_value(source_sheet.cell(row=row, column=4))
            logger.debug("WBS type for row {}: {}", row, wbs_type)
            
            startup_days = self.startup_analyzer.analyze_startup_days(
                formula,
                self.primitive_formulas,
                self.primitive_data,
                row,
                source_sheet,
                source_formulas,
                self.filename,
                element_type,
                wbs_type
            )
            
            if startup_days is not None:
                logger.info("Got startup days: {}", startup_days)
                target_sheet.cell(row=self.current_output_row, column=5).value = startup_days
            else:
                logger.warning("No startup days returned from analyzer")
                
        except Exception as e:
            logger.error("Error analyzing startup days for row {}: {}", row, str(e))
    
    def process(self, source_sheet: openpyxl.worksheet.worksheet.Worksheet, target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                source_formulas: Optional[openpyxl.worksheet.worksheet.Worksheet] = None) -> bool:
        """Process SCHEMA sheet with new format."""
        try:
            logger.info("Processing SCHEMA sheet...")
            
            # Set up headers
            self.setup_headers(target_sheet)
            
            # Find header row in source
            from .structure_analyzer import find_header_row
            self.header_row = find_header_row(source_sheet)
            
            # Get PRIMITIVE sheets from workbook
            workbook = source_sheet.parent
            self.primitive_data = workbook["PRIMITIVE"] if "PRIMITIVE" in workbook else None
            self.primitive_formulas = workbook["PRIMITIVE"] if "PRIMITIVE" in workbook else None
            
            # Process rows
            for row in range(self.header_row + 1, source_sheet.max_row + 1):
                self.process_row(row, source_sheet, target_sheet, source_formulas)
                
            return True
            
        except Exception as e:
            logger.error("Failed to process SCHEMA sheet: {}", str(e))
            return False
