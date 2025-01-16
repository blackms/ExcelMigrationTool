"""Excel workbook handling utilities."""
import openpyxl
from typing import Optional
from dedaexcelai.logger import get_logger

logger = get_logger()

def load_workbooks(input_path: str) -> Optional[openpyxl.Workbook]:
    """Load workbooks from input path."""
    try:
        return openpyxl.load_workbook(input_path, data_only=False)
    except Exception as e:
        logger.error("Failed to load workbook: {}", str(e))
        return None

def create_output_workbook() -> openpyxl.Workbook:
    """Create new output workbook."""
    return openpyxl.Workbook()

__all__ = [
    'load_workbooks',
    'create_output_workbook'
]
