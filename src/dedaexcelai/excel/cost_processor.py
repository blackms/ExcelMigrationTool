from abc import ABC, abstractmethod
import openpyxl
from typing import Optional
from ..logger import get_logger
from .cell_operations import get_cell_value

logger = get_logger()

class CostProcessor(ABC):
    """Base class for processing different cost types."""
    
    def __init__(self, current_output_row: int):
        self.current_output_row = current_output_row
    
    def copy_value(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                  target_sheet: openpyxl.worksheet.worksheet.Worksheet, source_col: int, target_col: int) -> None:
        """Copy value from source to target cell."""
        value = get_cell_value(source_sheet.cell(row=row, column=source_col))
        if value is not None:
            target_cell = target_sheet.cell(row=self.current_output_row, column=target_col)
            target_cell.value = value
            # Set number format with 4 decimal places for numeric values
            if isinstance(value, (int, float)):
                target_cell.number_format = '#,##0.0000'
    
    @abstractmethod
    def process_columns(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                       target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                       element_type: str) -> None:
        """Process columns based on cost type."""
        pass

class FixedCostProcessor(CostProcessor):
    """Processor for Fixed costs (Optional/Mandatory)."""
    
    def process_columns(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                       target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                       element_type: str) -> None:
        """Process Fixed cost columns."""
        if element_type == 'SubElement':
            # Copy Fixed costs (H -> L, L -> N)
            self.copy_value(row, source_sheet, target_sheet, 8, 12)   # H -> L (Startup Costo)
            self.copy_value(row, source_sheet, target_sheet, 12, 14)  # L -> N (Startup Prezzo)
            
            # Set fixed margin 39.30% in column M
            target_sheet.cell(row=self.current_output_row, column=13).value = 0.3930  # M (Startup Margine)
            target_sheet.cell(row=self.current_output_row, column=13).number_format = '0.00%'
            
            logger.debug("Processed Fixed costs for SubElement - H{} -> L{}, L{} -> N{}, M{} = 39.30%", 
                        row, self.current_output_row, row, self.current_output_row, self.current_output_row)

class FeeCostProcessor(CostProcessor):
    """Processor for Fee costs (Optional/Mandatory)."""
    
    def process_columns(self, row: int, source_sheet: openpyxl.worksheet.worksheet.Worksheet,
                       target_sheet: openpyxl.worksheet.worksheet.Worksheet,
                       element_type: str) -> None:
        """Process Fee cost columns."""
        if element_type == 'SubElement':
            # Copy Fee costs (I -> O, M -> Q)
            self.copy_value(row, source_sheet, target_sheet, 9, 15)   # I -> O (Canone Costo)
            self.copy_value(row, source_sheet, target_sheet, 13, 17)  # M -> Q (Canone Prezzo)
            
            # Set fixed margin 39.30% in column P
            target_sheet.cell(row=self.current_output_row, column=16).value = 0.3930  # P (Canone Margine)
            target_sheet.cell(row=self.current_output_row, column=16).number_format = '0.00%'
            
            logger.debug("Processed Fee costs for SubElement - I{} -> O{}, M{} -> Q{}, P{} = 39.30%", 
                        row, self.current_output_row, row, self.current_output_row, self.current_output_row)

def create_cost_processor(cost_type: str, current_output_row: int) -> CostProcessor:
    """Factory function to create appropriate cost processor."""
    if 'Fixed' in cost_type:
        return FixedCostProcessor(current_output_row)
    elif 'Fee' in cost_type:
        return FeeCostProcessor(current_output_row)
    else:
        raise ValueError(f"Unknown cost type: {cost_type}")
