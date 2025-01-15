import openpyxl
from pathlib import Path
from typing import Dict, Any, Optional, Tuple
import re
from copy import copy
from decimal import Decimal
from .logger import get_logger

logger = get_logger()

def get_cell_value(cell: Optional[Any]) -> Any:
    """Helper function to safely get cell value"""
    if cell is None:
        return None
    return cell.value

def is_empty_or_dashes(value: Any) -> bool:
    """Check if a value is empty, None, or contains only dashes"""
    if value is None or value == "":
        return True
    if isinstance(value, str):
        return value.strip().replace("-", "") == ""
    return False

def is_number(value: Any) -> bool:
    """Check if a value is a number (including string representations)"""
    if value is None:
        return False
    try:
        Decimal(str(value))
        return True
    except:
        return False

def find_element_catalog_interval(sheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int) -> Tuple[int, int]:
    """
    Find the interval for the current element catalog by looking for separator rows
    Returns: (start_row, end_row)
    """
    # First, go back to find the start of this catalog if we're in the middle
    while start_row > 1:
        service_element = get_cell_value(sheet.cell(row=start_row - 1, column=2))
        if is_empty_or_dashes(service_element) and isinstance(service_element, str) and "-" in service_element:
            break
        start_row -= 1
    
    # Then find the end (next separator row)
    end_row = start_row
    max_row = sheet.max_row
    
    while end_row <= max_row:
        service_element = get_cell_value(sheet.cell(row=end_row, column=2))
        if end_row > start_row and is_empty_or_dashes(service_element) and isinstance(service_element, str) and "-" in service_element:
            end_row -= 1  # Don't include the separator row
            break
        end_row += 1
    
    if end_row > max_row:
        end_row = max_row
    
    return (start_row, end_row)

def determine_cost_type(row: int, sheet: openpyxl.worksheet.worksheet.Worksheet) -> str:
    """
    Determine the cost type based on WBS column within the element catalog interval
    Returns: "Fee Optional" for CANONE, "Fixed Optional" for FIXED
    """
    # Find the interval for the current element catalog
    interval_start, interval_end = find_element_catalog_interval(sheet, row)
    
    # Get the service element name for better logging
    service_element = get_cell_value(sheet.cell(row=interval_start, column=2))
    logger.debug(f"Element catalog '<yellow>{service_element}</yellow>' (rows <blue>{interval_start}</blue>-<blue>{interval_end}</blue>)")
    
    # Look for FIXED or CANONE in the WBS column (column D) within the interval
    wbs_type = None
    for r in range(interval_start, interval_end + 1):
        wbs_value = get_cell_value(sheet.cell(row=r, column=4))  # Column D is 4
        if isinstance(wbs_value, str):
            wbs_value = wbs_value.upper().strip()
            if wbs_value == "FIXED":
                wbs_type = "FIXED"
                logger.debug(f"Found <magenta>FIXED</magenta> type at row <blue>{r}</blue>")
                break
            elif wbs_value == "CANONE":
                wbs_type = "CANONE"
                logger.debug(f"Found <magenta>CANONE</magenta> type at row <blue>{r}</blue>")
                break
    
    if wbs_type == "FIXED":
        logger.debug(f"Using <green>Fixed Optional</green> based on WBS type")
        return "Fixed Optional"
    elif wbs_type == "CANONE":
        logger.debug(f"Using <green>Fee Optional</green> based on WBS type")
        return "Fee Optional"
    else:
        logger.warning(f"No WBS type found in interval, defaulting to <yellow>Fee Optional</yellow>")
        return "Fee Optional"

def find_header_row(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Find the row containing headers"""
    for row in range(1, sheet.max_row + 1):
        if get_cell_value(sheet.cell(row=row, column=1)) == "Portfolio":
            return row
    return 1

def copy_cell_style(source_cell: Any, target_cell: Any):
    """
    Safely copy cell style attributes individually to avoid StyleProxy issues
    """
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def set_euro_format(cell: Any):
    """Set cell format to Euro"""
    cell.number_format = '#,##0.00 €'

def migrate_excel(input_file: str, output_file: str, template_file: str) -> bool:
    """
    Migrate data from input Excel file to output Excel file using the specified template.
    
    Args:
        input_file: Path to the input Excel file
        output_file: Path where the output Excel file will be saved
        template_file: Path to the template Excel file
        
    Returns:
        bool: True if migration was successful, False otherwise
    """
    try:
        logger.info(f"Starting migration from <blue>{input_file}</blue> to <blue>{output_file}</blue>")
        logger.info(f"Using template: <blue>{template_file}</blue>")
        
        # Load workbooks without data_only to preserve formulas
        input_wb = openpyxl.load_workbook(input_file, data_only=True)  # Use data_only=True to get values from formulas
        template_wb = openpyxl.load_workbook(template_file)
        
        # Get SCHEMA sheet from input workbook
        if "SCHEMA" not in input_wb.sheetnames:
            logger.error("Input file does not contain a 'SCHEMA' sheet")
            raise ValueError("Input file does not contain a 'SCHEMA' sheet")
        
        input_sheet = input_wb["SCHEMA"]
        output_sheet = template_wb.active
        
        logger.info("<green>Found SCHEMA sheet, processing data...</green>")
        
        # Find header row in input sheet
        header_row = find_header_row(input_sheet)
        logger.debug(f"Header row found at row <blue>{header_row}</blue>")
        
        # Process each row after headers
        current_output_row = 2  # Start after headers in template
        rows_processed = 0
        
        for row in range(header_row + 1, input_sheet.max_row + 1):
            # Get service element from column B
            service_element = get_cell_value(input_sheet.cell(row=row, column=2))
            
            # Skip completely empty rows
            if service_element is None:
                continue
            
            # Handle empty or dash-only service elements differently
            if is_empty_or_dashes(service_element):
                logger.debug(f"Empty row <blue>{row}</blue>, adding formula to column P")
                # Add formula to column P (16): =N{row}/(1-O{row})
                cell = output_sheet.cell(row=current_output_row, column=16)
                formula = f"=N{current_output_row}/(1-O{current_output_row})"
                cell.value = formula
                set_euro_format(cell)
                logger.debug(f"Added formula: <yellow>{formula}</yellow> to row <blue>{current_output_row}</blue>")
            else:
                logger.debug(f"Processing row <blue>{row}</blue> -> <green>{current_output_row}</green>")
                
                # Copy product element (column B)
                output_sheet.cell(row=current_output_row, column=2).value = service_element
                
                # Determine and set cost type (column C)
                cost_type = determine_cost_type(row, input_sheet)
                output_sheet.cell(row=current_output_row, column=3).value = cost_type
                
                # Set Euro format for column P
                set_euro_format(output_sheet.cell(row=current_output_row, column=16))
            
            current_output_row += 1
            rows_processed += 1
        
        logger.info(f"Successfully processed <green>{rows_processed}</green> rows")
        
        # Save the output workbook
        template_wb.save(output_file)
        logger.success(f"Migration completed successfully. Output saved to: <blue>{output_file}</blue>")
        return True
        
    except Exception as e:
        logger.exception(f"Migration failed: <red>{str(e)}</red>")
        return False